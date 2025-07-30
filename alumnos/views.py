import xlrd
import openpyxl
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import Group
from django.db.models import Count, Q
from django.shortcuts import render, redirect
from django.utils.timezone import now, localtime
from datetime import date
from django.core.files.uploadedfile import InMemoryUploadedFile

from .models import Alumno, Curso
from atrasos.models import Atraso
from datetime import datetime

from django.db.models.functions import TruncDay

from django.http import HttpResponse

import io
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

import os
from datetime import datetime
from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from alumnos.models import Alumno, Curso
from django.conf import settings
import openpyxl
from django.utils.timezone import now

from django.shortcuts import render, get_object_or_404, redirect
from .forms import AlumnoForm
import unicodedata
from atrasos.models import Atraso  # ✅ Correcto si está definido en la app 'atrasos'



# ------------------------------
# 1. Vista HOME (Login)
# ------------------------------
def home(request):
    if request.user.is_authenticated:
        if request.user.groups.filter(name='administrador').exists():
            return redirect('dashboard_admin')
        else:
            return redirect('registrar_atraso')  # o a otra vista para inspectores

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        usuario = authenticate(request, username=username, password=password)
        if usuario is not None:
            login(request, usuario)
            if usuario.groups.filter(name='administrador').exists():
                return redirect('dashboard_admin')
            else:
                return redirect('registrar_atraso')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos.')

    return render(request, 'alumnos/home.html')


# ------------------------------
# 2. Carga de Excel SIGE
# ------------------------------


def cargar_excel_sige(request):
    if request.method == 'POST':
        archivo = request.FILES['archivo']
        extension = archivo.name.split('.')[-1].lower()

        try:
            if extension == 'xls':
                import xlrd
                libro = xlrd.open_workbook(file_contents=archivo.read())
                hoja = libro.sheet_by_index(0)
                encabezados = hoja.row_values(0)

                idx_nombres = encabezados.index("Nombres")
                idx_apellido_p = encabezados.index("Apellido Paterno")
                idx_apellido_m = encabezados.index("Apellido Materno")
                idx_grado = encabezados.index("Desc Grado")
                idx_letra = encabezados.index("Letra Curso")

                for i in range(1, hoja.nrows):
                    fila = hoja.row_values(i)

                    nombres = str(fila[idx_nombres]).strip() if fila[idx_nombres] else ''
                    apellido_p = str(fila[idx_apellido_p]).strip() if fila[idx_apellido_p] else ''
                    apellido_m = str(fila[idx_apellido_m]).strip() if fila[idx_apellido_m] else ''
                    grado = str(fila[idx_grado]).strip() if fila[idx_grado] else ''
                    letra = str(fila[idx_letra]).strip() if fila[idx_letra] else ''

                    # Requiere grado, letra y al menos un nombre o apellido
                    if not grado or not letra or not any([nombres, apellido_p, apellido_m]):
                        continue

                    curso_nombre = f"{grado} {letra}"
                    curso, _ = Curso.objects.get_or_create(nombre=curso_nombre)

                    nombre_completo = " ".join(filter(None, [apellido_p, apellido_m, nombres]))
                    Alumno.objects.get_or_create(nombre_completo=nombre_completo, curso=curso)

            elif extension == 'xlsx':
                import openpyxl
                wb = openpyxl.load_workbook(archivo)
                hoja = wb.active
                encabezados = [cell.value for cell in next(hoja.iter_rows(min_row=1, max_row=1))]

                idx_nombres = encabezados.index("Nombres")
                idx_apellido_p = encabezados.index("Apellido Paterno")
                idx_apellido_m = encabezados.index("Apellido Materno")
                idx_grado = encabezados.index("Desc Grado")
                idx_letra = encabezados.index("Letra Curso")

                for fila in hoja.iter_rows(min_row=2):
                    nombres = str(fila[idx_nombres].value).strip() if fila[idx_nombres].value else ''
                    apellido_p = str(fila[idx_apellido_p].value).strip() if fila[idx_apellido_p].value else ''
                    apellido_m = str(fila[idx_apellido_m].value).strip() if fila[idx_apellido_m].value else ''
                    grado = str(fila[idx_grado].value).strip() if fila[idx_grado].value else ''
                    letra = str(fila[idx_letra].value).strip() if fila[idx_letra].value else ''

                    if not grado or not letra or not any([nombres, apellido_p, apellido_m]):
                        continue

                    curso_nombre = f"{grado} {letra}"
                    curso, _ = Curso.objects.get_or_create(nombre=curso_nombre)

                    nombre_completo = " ".join(filter(None, [apellido_p, apellido_m, nombres]))
                    Alumno.objects.get_or_create(nombre_completo=nombre_completo, curso=curso)

            else:
                messages.error(request, "Formato no válido. Usa archivos .xls o .xlsx.")
                return redirect('cargar_excel')

        except Exception as e:
            messages.error(request, f"Ocurrió un error al procesar el archivo: {e}")
            return redirect('cargar_excel')

        messages.success(request, "Alumnos cargados correctamente desde el SIGE.")
        return redirect('dashboard_admin')

    return render(request, 'alumnos/cargar_excel.html')




# ------------------------------
# 3. Panel de administración
# ------------------------------
@login_required
def dashboard_admin(request):
    if not request.user.groups.filter(name='administrador').exists():
        return redirect('home')

    hoy = now().date()
    mes = hoy.month
    anio = hoy.year

    cursos_disponibles = Curso.objects.all()
    alumnos_disponibles = Alumno.objects.all()

    # Captura de filtros
    curso_id = request.GET.get('curso')
    alumno_id = request.GET.get('alumno')
    fecha_filtro = request.GET.get('fecha')  # ← Nuevo filtro de fecha
    fecha_dt = None

    if fecha_filtro:
        try:
            fecha_dt = datetime.strptime(fecha_filtro, '%Y-%m-%d').date()
        except ValueError:
            fecha_dt = None

    atrasos_mes = Atraso.objects.filter(fecha__month=mes, fecha__year=anio)

    # Aplicar filtros adicionales
    if curso_id:
        atrasos_mes = atrasos_mes.filter(alumno__curso_id=curso_id)
    if alumno_id:
        atrasos_mes = atrasos_mes.filter(alumno_id=alumno_id)
    if fecha_dt:
        atrasos_mes = atrasos_mes.filter(fecha=fecha_dt)

    total_alumnos = Alumno.objects.count()
    atrasos_hoy = Atraso.objects.filter(fecha=hoy).count()
    promedio_mensual = round(atrasos_mes.count() / total_alumnos, 1) if total_alumnos else 0

    atrasos_por_curso = (
        atrasos_mes
        .values('alumno__curso__nombre')
        .annotate(total=Count('id'))
        .order_by('alumno__curso__nombre')
    )

    alumnos_con_mas_atrasos = (
        atrasos_mes
        .values('alumno__nombre_completo', 'alumno__curso__nombre')
        .annotate(total=Count('id'))
        .order_by('-total')[:5]
    )

    alertas = (
        atrasos_mes
        .values('alumno__nombre_completo', 'alumno__curso__nombre')
        .annotate(total=Count('id'))
        .filter(total__gte=3)
        .order_by('-total')
    )

    # Alumnos para tabla final
    alumnos = Alumno.objects.select_related('curso').all()
    if curso_id:
        alumnos = alumnos.filter(curso_id=curso_id)
    if alumno_id:
        alumnos = alumnos.filter(id=alumno_id)
    alumnos = alumnos.order_by('curso__nombre', 'nombre_completo')

    context = {
        'total_alumnos': total_alumnos,
        'atrasos_hoy': atrasos_hoy,
        'promedio_mensual': promedio_mensual,
        'atrasos_por_curso': atrasos_por_curso,
        'alumnos_con_mas_atrasos': alumnos_con_mas_atrasos,
        'alertas': alertas,
        'alumnos': alumnos,
        'cursos_disponibles': cursos_disponibles,
        'alumnos_disponibles': alumnos_disponibles,
        'fecha_seleccionada': fecha_filtro  # ← Lo pasamos a la plantilla
    }

    return render(request, 'alumnos/dashboard_admin.html', context)


# ------------------------------
# 4. Vista de alertas (3 atrasos)
# ------------------------------
@login_required
def alertas_atrasos(request):
    if not request.user.groups.filter(name='administrador').exists():
        return redirect('home')

    hoy = now().date()
    mes = hoy.month
    anio = hoy.year

    alertas = (
        Atraso.objects
        .filter(fecha__month=mes, fecha__year=anio)
        .values('alumno', 'alumno__nombre_completo', 'alumno__curso__nombre')
        .annotate(total=Count('id'))
        .filter(total__gte=3)
        .order_by('-total')
    )

    return render(request, 'alumnos/alertas_atrasos.html', {'alertas': alertas})


# ------------------------------
# (Opcional) Vista lista alumnos clásica
# ------------------------------
@login_required
def lista_alumnos(request):
    alumnos = Alumno.objects.all().order_by('nombre_completo')
    return render(request, 'alumnos/lista_alumnos.html', {'alumnos': alumnos})


from collections import defaultdict
from django.db.models import Count, Q
from django.db.models.functions import TruncDay

@login_required
def estadisticas(request):
    hoy = now().date()
    mes = hoy.month
    anio = hoy.year
    atrasos_mes = Atraso.objects.filter(fecha__month=mes, fecha__year=anio)

    # 1. Gráfico de líneas - Atrasos por día
    atrasos_por_dia = (
        atrasos_mes
        .annotate(dia=TruncDay('fecha'))
        .values('dia')
        .annotate(total=Count('id'))
        .order_by('dia')
    )
    dias_labels = [a['dia'].strftime('%d-%m') for a in atrasos_por_dia]
    dias_values = [a['total'] for a in atrasos_por_dia]

    # 2. Gráfico de pastel - alumnos con 3 o más atrasos
    total_alumnos = Alumno.objects.count()
    con_3_mas = (
        atrasos_mes
        .values('alumno')
        .annotate(total=Count('id'))
        .filter(total__gte=3)
        .count()
    )
    otros = total_alumnos - con_3_mas
    pie_values = [con_3_mas, max(otros, 0)]

    # 3. Gráfico apilado - alumnos con menos de 3 y con 3 o más atrasos por curso
    alumnos_con_atrasos = (
        atrasos_mes
        .values('alumno__id', 'alumno__curso__nombre')
        .annotate(total=Count('id'))
    )

    stacked_labels = sorted(set([x['alumno__curso__nombre'] for x in alumnos_con_atrasos]))
    stacked_menos_3 = []
    stacked_3_mas = []

    for curso in stacked_labels:
        menos_3 = sum(1 for a in alumnos_con_atrasos if a['alumno__curso__nombre'] == curso and a['total'] < 3)
        mas_3 = sum(1 for a in alumnos_con_atrasos if a['alumno__curso__nombre'] == curso and a['total'] >= 3)
        stacked_menos_3.append(menos_3)
        stacked_3_mas.append(mas_3)

    # 4. Gráfico - Total atrasos vs Justificados vs Total real
    cursos = Curso.objects.all()
    grafico_cursos = []
    grafico_justificados = []
    grafico_reales = []

    for curso in cursos:
        total = atrasos_mes.filter(alumno__curso=curso).count()
        justificados = atrasos_mes.filter(alumno__curso=curso, estado='justificado').count()
        real = total - justificados
        grafico_cursos.append(curso.nombre)
        grafico_justificados.append(justificados)
        grafico_reales.append(real)

    # 5. Cursos disponibles y gráfico de alumnos filtrado por curso
    cursos_disponibles = Curso.objects.all()
    curso_actual = request.GET.get('curso')

    if curso_actual:
        try:
            curso_obj = Curso.objects.get(id=int(curso_actual))
        except (ValueError, Curso.DoesNotExist):
            curso_obj = None
    else:
        # Buscar el primer curso con atrasos en el mes si no se selecciona ninguno
        cursos_con_datos_ids = Atraso.objects.filter(
            fecha__month=mes,
            fecha__year=anio
        ).values_list('alumno__curso__id', flat=True).distinct()

        cursos_con_datos = Curso.objects.filter(id__in=cursos_con_datos_ids)
        curso_obj = cursos_con_datos.first() if cursos_con_datos.exists() else None

    # Obtener alumnos filtrados
    alumnos_filtrados = Alumno.objects.filter(curso=curso_obj) if curso_obj else Alumno.objects.none()

    # 6. Datos por alumno (Total y Justificados)
    datos_por_alumno = []
    for alumno in alumnos_filtrados:
        total = atrasos_mes.filter(alumno=alumno).count()
        justificados = atrasos_mes.filter(alumno=alumno, estado='justificado').count()
        if total > 0:
            datos_por_alumno.append({
                'nombre': alumno.nombre_completo,
                'total': total,
                'justificados': justificados,
            })

    nombres_alumnos = [d['nombre'] for d in datos_por_alumno]
    atrasos_totales = [d['total'] for d in datos_por_alumno]
    atrasos_justificados = [d['justificados'] for d in datos_por_alumno]

    # 7. Enviar todo al template
    context = {
        'dias_labels': dias_labels,
        'dias_values': dias_values,
        'pie_values': pie_values,
        'stacked_labels': stacked_labels,
        'stacked_menos_3': stacked_menos_3,
        'stacked_3_mas': stacked_3_mas,
        'grafico_cursos': grafico_cursos,
        'grafico_justificados': grafico_justificados,
        'grafico_reales': grafico_reales,
        'cursos_disponibles': cursos_disponibles,
        'curso_actual': curso_obj.id if curso_obj else '',
        'nombres_alumnos': nombres_alumnos,
        'atrasos_totales': atrasos_totales,
        'atrasos_justificados': atrasos_justificados,
    }

    return render(request, 'alumnos/estadisticas.html', context)


@login_required
def exportar_excel(request):
    atrasos = Atraso.objects.select_related('alumno__curso', 'registrado_por').order_by('fecha')

    # Preconteo de atrasos por alumno
    atrasos_por_alumno = defaultdict(int)
    for atraso in atrasos:
        atrasos_por_alumno[atraso.alumno_id] += 1

    # Agrupar atrasos por mes
    datos_por_mes = defaultdict(list)
    for atraso in atrasos:
        mes = atraso.fecha.strftime("%B %Y")  # Ej: "Julio 2025"
        datos_por_mes[mes].append(atraso)

    # Crear libro Excel
    wb = Workbook()
    ws_inicio = wb.active
    ws_inicio.title = "Resumen General"

    columnas = ["Alumno", "Curso", "Fecha", "Hora Llegada", "Minutos Tarde", "Estado", "Registrado por"]
    ws_inicio.append(columnas)

    # Estilos
    encabezado_font = Font(bold=True)
    color_amarillo = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

    for celda in ws_inicio[1]:
        celda.font = encabezado_font

    curso_contador = defaultdict(int)

    # Agregar datos
    for atraso in atrasos:
        nombre_usuario = "Desconocido"
        if atraso.registrado_por:
            nombre_usuario = atraso.registrado_por.get_full_name() or atraso.registrado_por.username

        fila = [
            atraso.alumno.nombre_completo,
            atraso.alumno.curso.nombre,
            atraso.fecha.strftime('%d-%m-%Y'),
            atraso.hora_llegada.strftime('%H:%M'),
            atraso.minutos_tarde,
            atraso.estado,
            nombre_usuario
        ]
        ws_inicio.append(fila)
        curso_contador[atraso.alumno.curso.nombre] += 1

        # Pintar en amarillo si tiene 3 o más atrasos
        if atrasos_por_alumno[atraso.alumno_id] >= 3:
            fila_idx = ws_inicio.max_row
            for celda in ws_inicio[f"A{fila_idx}:G{fila_idx}"][0]:
                celda.fill = color_amarillo

    # Ajustes visuales de la hoja principal
    if ws_inicio.max_row > 1:
        ws_inicio.auto_filter.ref = f"A1:G{ws_inicio.max_row}"
    for col in range(1, 8):
        ws_inicio.column_dimensions[get_column_letter(col)].width = 20

    # ✅ Crear hojas por mes
    for mes, atrasos_mes in datos_por_mes.items():
        ws_mes = wb.create_sheet(title=mes)
        ws_mes.append(columnas)
        for celda in ws_mes[1]:
            celda.font = encabezado_font

        for atraso in atrasos_mes:
            nombre_usuario = "Desconocido"
            if atraso.registrado_por:
                nombre_usuario = atraso.registrado_por.get_full_name() or atraso.registrado_por.username

            ws_mes.append([
                atraso.alumno.nombre_completo,
                atraso.alumno.curso.nombre,
                atraso.fecha.strftime('%d-%m-%Y'),
                atraso.hora_llegada.strftime('%H:%M'),
                atraso.minutos_tarde,
                atraso.estado,
                nombre_usuario
            ])

        if ws_mes.max_row > 1:
            ws_mes.auto_filter.ref = f"A1:G{ws_mes.max_row}"
        for col in range(1, 8):
            ws_mes.column_dimensions[get_column_letter(col)].width = 20

    # ✅ Hoja de estadísticas con gráfico
    ws_grafico = wb.create_sheet("Estadísticas")
    ws_grafico.append(["Curso", "Total Atrasos"])
    for celda in ws_grafico[1]:
        celda.font = encabezado_font

    for curso, total in curso_contador.items():
        ws_grafico.append([curso, total])

    if curso_contador:
        chart = BarChart()
        chart.title = "Atrasos por Curso"
        chart.y_axis.title = "Cantidad de Atrasos"
        chart.x_axis.title = "Curso"
        chart.style = 2

        data = Reference(ws_grafico, min_col=2, min_row=1, max_row=len(curso_contador)+1)
        cats = Reference(ws_grafico, min_col=1, min_row=2, max_row=len(curso_contador)+1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        ws_grafico.add_chart(chart, "D2")

    # Descargar el archivo
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename=Informe_Atrasos_Completo.xlsx'
    return response


@login_required
def opciones_admin(request):
    return render(request, 'alumnos/opciones_admin.html')

from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from datetime import datetime
from .models import Alumno, Curso

# Validar si es administrador
def es_admin(user):
    return user.is_superuser or user.groups.filter(name='admin').exists()


@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='admin').exists())
def iniciar_nuevo_anio(request):
    anio_actual = datetime.now().year
    lista_anios = list(range(anio_actual, anio_actual - 6, -1))

    if request.method == "POST":
        accion = request.POST.get("accion")

        if accion == "cambiar_anio":
            anio_seleccionado = int(request.POST.get("anio_seleccionado"))
            request.session["anio_trabajo"] = anio_seleccionado
            messages.success(request, f"Año de trabajo cambiado a {anio_seleccionado}.")
            return redirect("iniciar_nuevo_anio")

        elif accion == "restaurar":
            anio_restaurar = int(request.POST.get("anio_restaurar"))
            ruta_archivo = os.path.join(settings.MEDIA_ROOT, f"respaldos/alumnos_{anio_restaurar}.xlsx")

            if os.path.exists(ruta_archivo):
                wb = openpyxl.load_workbook(ruta_archivo)
                ws = wb.active
                Alumno.objects.all().delete()

                for row in ws.iter_rows(min_row=2, values_only=True):
                    nombre, curso_nombre = row
                    curso, _ = Curso.objects.get_or_create(nombre=curso_nombre)
                    Alumno.objects.create(nombre_completo=nombre, curso=curso)

                messages.success(request, f"Datos del año {anio_restaurar} restaurados correctamente.")
            else:
                messages.error(request, f"No se encontró respaldo para el año {anio_restaurar}.")
            return redirect("iniciar_nuevo_anio")

        elif accion == "iniciar_nuevo":
            nombre_archivo = f"alumnos_{anio_actual}.xlsx"
            carpeta_respaldo = os.path.join(settings.MEDIA_ROOT, "respaldos")
            os.makedirs(carpeta_respaldo, exist_ok=True)
            ruta_completa = os.path.join(carpeta_respaldo, nombre_archivo)

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.append(["Nombre", "Curso"])

            for alumno in Alumno.objects.all():
                ws.append([alumno.nombre_completo, alumno.curso.nombre])

            wb.save(ruta_completa)

            egresados = Alumno.objects.filter(curso__nombre__startswith="8° básico")
            egresados.delete()

            for alumno in Alumno.objects.exclude(curso__nombre__startswith="8° básico"):
                nombre_curso = alumno.curso.nombre.lower()

                if "1er nivel de transicion" in nombre_curso:
                    nuevo_nombre = alumno.curso.nombre.replace("1er", "2do")
                elif "2do nivel de transicion" in nombre_curso:
                    nuevo_nombre = alumno.curso.nombre
                else:
                    try:
                        numero = int(nombre_curso[0])
                        letra = alumno.curso.nombre[-1]
                        nuevo_nombre = f"{numero + 1}° básico {letra}"
                    except:
                        continue

                nuevo_curso, _ = Curso.objects.get_or_create(nombre=nuevo_nombre)
                alumno.curso = nuevo_curso
                alumno.save()

            messages.success(request, f"Año escolar {anio_actual + 1} iniciado correctamente. Respaldo guardado.")
            return redirect("iniciar_nuevo_anio")

    return render(request, "alumnos/iniciar_nuevo_anio.html", {
        "lista_anios": lista_anios,
        "anio_actual": request.session.get("anio_trabajo", anio_actual)
    })



import unicodedata
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.shortcuts import render
from .models import Alumno, Curso

# Función para normalizar tildes
def normalizar(texto):
    return unicodedata.normalize('NFKD', texto).encode('ASCII', 'ignore').decode('ASCII').lower()

@login_required
def lista_alumnos(request):
    alumnos = Alumno.objects.select_related('curso').all()

    nombre_filtro = request.GET.get('nombre', '').strip()
    curso_filtro = request.GET.get('curso', '').strip()

    if nombre_filtro:
        alumnos = alumnos.filter(nombre_completo__icontains=nombre_filtro)
    if curso_filtro:
        alumnos = alumnos.filter(curso__id=curso_filtro)

    # Ordenar alfabéticamente ignorando tildes
    alumnos = sorted(alumnos, key=lambda a: normalizar(a.nombre_completo))

    # Paginación: 20 alumnos por página
    paginator = Paginator(alumnos, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    cursos = Curso.objects.all().order_by('nombre')

    return render(request, 'alumnos/lista_alumnos.html', {
        'page_obj': page_obj,
        'cursos': cursos,
        'request': request  # para conservar filtros en la plantilla
    })


@login_required
def agregar_alumno(request):
    if request.method == 'POST':
        form = AlumnoForm(request.POST)
        if form.is_valid():
            alumno = form.save(commit=False)
            alumno.nombre_completo = alumno.nombre_completo.upper()  # Guardar en mayúsculas
            alumno.save()
            messages.success(request, "Alumno agregado correctamente.")
            return redirect('lista_alumnos')
    else:
        form = AlumnoForm()
    return render(request, 'alumnos/agregar_alumno.html', {'form': form})

@login_required
def editar_alumno(request, alumno_id):
    alumno = get_object_or_404(Alumno, id=alumno_id)
    if request.method == 'POST':
        form = AlumnoForm(request.POST, instance=alumno)
        if form.is_valid():
            form.save()
            messages.success(request, "Alumno actualizado.")
            return redirect('lista_alumnos')
    else:
        form = AlumnoForm(instance=alumno)
    return render(request, 'alumnos/editar_alumno.html', {'form': form})



@login_required
def confirmar_eliminar_alumno(request, alumno_id):
    alumno = get_object_or_404(Alumno, id=alumno_id)

    if request.method == 'POST':
        alumno.delete()
        messages.success(request, "Alumno eliminado correctamente.")
        return redirect('lista_alumnos')

    return render(request, 'alumnos/confirmar_eliminar_alumno.html', {'alumno': alumno})



@login_required
@user_passes_test(lambda u: u.is_superuser or u.groups.filter(name='admin').exists())
def eliminar_toda_la_base(request):
    if request.method == 'POST':
        Alumno.objects.all().delete()
        Atraso.objects.all().delete()
        messages.success(request, "✅ Toda la base de datos ha sido eliminada correctamente.")
        return redirect('opciones_admin')
    return redirect('opciones_admin')

@login_required
@user_passes_test(lambda u: u.groups.filter(name='administrador').exists())
def eliminar_base(request):
    if request.method == 'POST':
        Atraso.objects.all().delete()
        Alumno.objects.all().delete()
        Curso.objects.all().delete()  # Asegúrate de limpiar también los cursos
        messages.success(request, "✅ Todos los datos han sido eliminados correctamente.")
        return redirect('opciones_admin')  # Redirige a Más Opciones
    else:
        messages.error(request, "⚠️ Acción no permitida.")
        return redirect('opciones_admin')
